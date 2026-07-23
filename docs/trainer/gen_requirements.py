# -*- coding: utf-8 -*-
"""Hyper-Parallel 需求表单一数据源生成器（端到端训练流程版）。

修改 ROWS/FLOW/MODELS 后运行 `python docs/gen_requirements.py`，
同时重新生成 requirements.xlsx / requirements.md / requirements.csv，保证三者不漂移。

2026-07-22 更新：每条需求新增第 9 个元素「状态」（已实现/部分实现/未实现），
按 hyper_models/components/distributed/ 与 core/dtensor 实现实况标注
（282 个测试全绿；设计文档 01-06 于 2026-07-21/22 完成两轮全面修订）。
"""

import csv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────────────────────────────────
# 需求数据：(阶段, 模块, 需求点, 需求描述, 适用模型, 工时, 优先级, 设计文档, 状态)
# 适用模型: 全部 / DeepSeek V3·V4 / Qwen3.5 / GLM-Image / Pangu(自研)
# 状态: 已实现 / 部分实现 / 未实现（按 components/distributed + core/dtensor 实况）
# ────────────────────────────────────────────────────────────────────────────

ALL = "全部"
DS = "DeepSeek V3/V4"
QW = "Qwen3.5"
GLM = "GLM-Image"
PG = "Pangu(自研)"

DONE = "已实现"
PART = "部分实现"
TODO = "未实现"

ROWS = [
    # ── S0 前置依赖 ──────────────────────────────────────────────────────
    ("S0 前置依赖", "外部依赖(FSDP2扩展)", "TP 梯度同步集成落地(D-12 二选一)",
     "D-12 裁决的集成落地二选一：(a) 调时序使 fully_shard 先于 _local_params_context 解包执行，"
     "走已实现的 DTENSOR_UNIFIED 路径（_orig_dtensor_placements + layout-driven 归约组），无需 fork；"
     "(b) 补齐 fully_shard(tp_grad_info=...) 消费端（非 stock API，需 fork 或上游 PR）。"
     "现状：tp_grad_info 由 planner/applier 产出但全仓库无消费者（仅 grad_equiv.py 测试模拟旁路）",
     ALL, 2, "P0", "05 §6.7.1/§7/§12.2 D-12", TODO),
    ("S0 前置依赖", "外部依赖(FSDP2扩展)", "_get_base_spmd_placements DTENSOR_UNIFIED 路径",
     "stock FSDP2 已实现 DTENSOR_UNIFIED 机制（无需 fork 新分支）：_get_base_spmd_placements "
     "经 DeviceMesh.concatenate 拼统一 mesh，_build_layout_driven_group_info 从最终 layout 的 "
     "Replicate 轴推 TP 归约组，_normalize_unsharded_grad_to_local 处理 Partial 梯度回流；"
     "剩余为集成时序接线（D-12 选项 a/b）",
     ALL, 1, "P0", "05 §6.7.2/§6.7.3", PART),
    ("S0 前置依赖", "外部依赖(FSDP2扩展)", "DeviceMesh.concatenate helper",
     "已实现：core/dtensor/device_mesh.py:1035（layout-backed 拼合 + 非重叠校验 + 进程组元数据继承），"
     "供 FSDP2 DTENSOR_UNIFIED 路径拼合 [dp,tp] 统一 mesh",
     ALL, 0.5, "P0", "05 §6.7.2", DONE),
    ("S0 前置依赖", "外部依赖(版本钉板)", "依赖版本钉板与 API 核验",
     "钉死并核验 torch/transformers/torchao/peft/accelerate/datasets/torchdata 版本，"
     "消除历史 PyTorch API 误用风险",
     ALL, 1, "P0", "—", TODO),
    ("S0 前置依赖", "外部依赖(模型可获得性)", "交付模型权重/架构可获得性确认",
     "确认 DeepSeek V3/V4、Qwen3.5、GLM-Image 在 HuggingFace 的权重与架构字符串来源；"
     "Pangu 权重与词表由内部渠道提供，确认交付形式",
     f"{DS} / {QW} / {GLM} / {PG}", 0.5, "P0", "—", TODO),

    # ── S1 启动与配置 ────────────────────────────────────────────────────
    ("S1 启动与配置", "配置解析", "YAML 强类型解析",
     "读取训练 YAML，经 resolve_root() 解析为强类型 TrainerConfig（拒绝未知一级字段）",
     ALL, 1, "P0", "01 §2", DONE),
    ("S1 启动与配置", "配置解析", "组件引用即时解析",
     "_target_ 字符串 → 实际 Python 类/函数（import_target 点分隔导入，类 target 注解取自 __init__）",
     ALL, 0.5, "P0", "01 §2.4", DONE),
    ("S1 启动与配置", "配置解析", "typed 值校验与 CLI override",
     "coerce_value 基于类型注解做校验与转换；CLI --field=value 走 typed 路径替换",
     ALL, 0.5, "P0", "01 §2.4/§2.5", DONE),
    ("S1 启动与配置", "配置解析", "组件即时构造",
     "resolve_component：签名校验 + typed 参数转换后立即调用 _target_ 构造 Config 对象",
     ALL, 1, "P0", "01 §2.4", DONE),
    ("S1 启动与配置", "配置解析", "辅助访问与序列化",
     "Configurable.Config 的 replace / to_dict / traverse（instantiate_path 与敏感字段脱敏未实现）",
     ALL, 0.5, "P1", "01 §3", "部分实现"),
    ("S1 启动与配置", "配置解析", "导入安全控制",
     "白名单前缀限制 + 用户模块开关，防止 YAML 执行未授权导入（原 01 §2.12 已移除，需求保留）",
     ALL, 0.5, "P1", "—", TODO),
    ("S1 启动与配置", "配置解析", "强类型配置协议",
     "TrainerConfig + Configurable.Config（build/replace/to_dict/traverse）：组件 typed 配置与构建协议",
     ALL, 1, "P0", "01 §3", DONE),

    # ── S2 分布式环境初始化 ─────────────────────────────────────────────
    ("S2 分布式环境", "并行拓扑", "并行度声明",
     "ParallelismSizes dataclass：用户声明 tp/cp/pp/dp/ep 各维并行度",
     ALL, 0.5, "P0", "06 §3.1", TODO),
    ("S2 分布式环境", "并行拓扑", "DeviceMesh 构建",
     "MeshContext.build()：按并行度构建多维 DeviceMesh，自动推导 dp_size；主 mesh 无 EP 轴"
     "（D-10：expert mesh 于 apply 期由全 dense 区域派生 (edp,ep)）；DTensor 管理 tp/cp，"
     "DP 由 FSDP2 管理",
     ALL, 1, "P0", "06 §3.2", TODO),
    ("S2 分布式环境", "并行拓扑", "网格查询接口",
     "MeshContext properties：各维 size/rank 实时查询",
     ALL, 0.5, "P0", "06 §3.2", TODO),
    ("S2 分布式环境", "并行拓扑", "统一拓扑容器",
     "DistributedSetup.build()：整合 MeshContext + strategy_config + pipeline_config 单一入口",
     ALL, 0.5, "P0", "06 §3.3", TODO),
    ("S2 分布式环境", "数据并行", "FSDP2 策略配置",
     "FSDP2Config：sequence_parallel/activation_checkpointing/mp_policy/offload_policy",
     ALL, 0.5, "P0", "06 §4.1", TODO),
    ("S2 分布式环境", "数据并行", "FSDP2 数据并行包裹",
     "FSDP2Manager.parallelize()：DTensor 分片完成后对每个参数施加 fully_shard",
     ALL, 1, "P0", "06 §4.2", TODO),
    ("S2 分布式环境", "数据并行", "参数本地化解包",
     "已实现：sharding/apply.py 的 _local_params_context 在 FSDP2 包裹前将 DTensor 参数"
     "零拷贝替换为 plain local tensor（与 fully_shard 的先后时序属 D-12 集成项）",
     ALL, 0.5, "P0", "05 §4.4 / 06 §5", DONE),
    ("S2 分布式环境", "数据并行", "TP 梯度同步元数据",
     "tp_grad.py build_tp_grad_info(plan, tp_mesh) 已实现：从 ShardingPlan 读取 TP placement；"
     "fully_shard 消费端接线待落地（D-12，全仓库暂无消费者，仅 grad_equiv.py 测试模拟旁路）",
     ALL, 1, "P0", "05 §6.7.1", PART),

    # ── S3 数据供给 ─────────────────────────────────────────────────────
    ("S3 数据供给", "数据加载", "数据加载器主流程",
     "串联 tokenizer → dataset → 分片 → packing → sampler → collate → DataLoader（7 步）",
     ALL, 1, "P0", "02", TODO),
    ("S3 数据供给", "数据加载", "分词器构建",
     "4 路分发（无 key / null / 无 _target_ / 有 _target_）加载 tokenizer",
     ALL, 0.5, "P0", "02", TODO),
    ("S3 数据供给", "数据加载", "HF 数据集加载",
     "datasets.load_dataset()，仅首个 rank 下载",
     ALL, 0.5, "P0", "02", TODO),
    ("S3 数据供给", "数据加载", "Megatron 数据集兼容",
     "兼容 Megatron 预处理后的二进制数据集格式",
     ALL, 0.5, "P1", "02", TODO),
    ("S3 数据供给", "序列打包", "文本序列打包",
     "多短序列拼接为长序列（THD 格式 + bin-packing 策略），提升 GPU 利用率",
     ALL, 0.5, "P0", "02", TODO),
    ("S3 数据供给", "序列打包", "图像打包(NEAT)",
     "GLM-Image 的 NEAT 图像打包：变分辨率图像 patch 序列化与文本交错打包",
     GLM, 1, "P0", "02", TODO),
    ("S3 数据供给", "采样与拼接", "分布式采样",
     "每个 DP rank 获取不重叠数据分片，支持断点续训状态恢复",
     ALL, 0.5, "P0", "02", TODO),
    ("S3 数据供给", "采样与拼接", "批次拼接",
     "per-key padding + labels 对齐 + CP 序列边界对齐 + PP 链式包装；"
     "其中 CP 边界对齐 shard_batch_for_cp 已在 cp_utils.py 实现（含 seq_lens 分片），"
     "其余环节未实现",
     ALL, 0.5, "P0", "02 / 05 §6.3.4", PART),

    # ── S4 模型构建与初始化 ─────────────────────────────────────────────
    ("S4 模型构建", "模型注册", "模型架构注册表",
     "架构名 → 模型类映射（懒加载 import）；HF 模型与自研模型统一注册",
     ALL, 0.5, "P0", "01 §5", TODO),
    ("S4 模型构建", "模型注册", "模型类型判定",
     "查表判别 HF 原生模型 vs 内置自研模型（Pangu 走内置路径）",
     f"{ALL}（{PG} 内置）", 0.5, "P0", "01 §5", TODO),
    ("S4 模型构建", "HF兼容层", "HF 标准接口兼容",
     "多重继承 HF AutoModel，提供 from_pretrained / from_config / _build_model",
     f"{DS} / {QW} / {GLM}", 1.5, "P0", "01 §6.1", TODO),
    ("S4 模型构建", "HF兼容层", "多任务类族",
     "ForCausalLM（DeepSeek/Qwen3.5/Pangu）+ ForImageTextToText（GLM-Image）",
     ALL, 0.5, "P0", "01 §6.1", TODO),
    ("S4 模型构建", "HF兼容层", "from_pretrained 完整实现",
     "distributed_setup / torch_dtype / attn_implementation / validate_placement 等参数",
     ALL, 1, "P0", "01 §6.2", TODO),
    ("S4 模型构建", "构建编排", "模型构建全流程编排",
     "meta device → PEFT → freeze → ShardingPlanner → ShardingApplier → 权重加载 → FSDP2，"
     "严格按依赖顺序（fully_shard 先于 to_empty/load，避免二次显存峰值）",
     ALL, 1.5, "P0", "01 §6.3 / 06 §5.2", TODO),
    ("S4 模型构建", "构建编排", "模型空壳创建",
     "meta device 实例化（零显存，仅记录 shape/dtype），自研/HF 路径分发",
     ALL, 1, "P0", "01 §7", TODO),
    ("S4 模型构建", "参数处理", "PEFT 注入",
     "分片前插入 LoRA 等适配器层",
     ALL, 0.5, "P1", "01 §6.4", TODO),
    ("S4 模型构建", "参数处理", "参数冻结",
     "按配置冻结指定参数（embedding、前 N 层等）",
     ALL, 0.5, "P1", "01 §6.5", TODO),

    # ── S5 并行分片引擎 ─────────────────────────────────────────────────
    ("S5 并行分片", "分片规划", "分片数据结构",
     "已实现：sharding_config.py 的 NamedPlacement / ModuleShardingSpec / ShardingPlan："
     "参数分片 + in/out 四元 placement 契约",
     ALL, 1.5, "P0", "05 §3.1-3.2", DONE),
    ("S5 并行分片", "分片规划", "分片模板库",
     "已实现：7 种模块类型（attention/mlp/norm/embed/lm_head/moe_gate/moe_mlp）× SP/non-SP，"
     "TP+CP+EP 三维 placement 声明（sharding_planner.TEMPLATES）",
     ALL, 1, "P0", "05 §3.5", DONE),
    ("S5 并行分片", "分片规划", "参数角色分类(Phase1)",
     "已实现：命名规则 + ARCH_OVERRIDES 架构覆盖 → 14 种 ParamRole"
     "（D-13 新增 REPLICATED：全维 Replicate，仅经 ARCH_OVERRIDES 显式指派，默认规则不产生）",
     ALL, 1.5, "P0", "05 §3.6.6/§12.2 D-13", DONE),
    ("S5 并行分片", "分片规划", "边界分组(Phase2)",
     "已实现：参数按通信边界模块聚合（两趟实现：直属分组 + 深度优先向上合并）",
     ALL, 0.5, "P0", "05 §12.3", DONE),
    ("S5 并行分片", "分片规划", "语义推断(Phase3)",
     "已实现：FQN 显式模式 > 叶守卫 > 参数角色组合 → boundary_type",
     ALL, 0.5, "P0", "05 §3.6.2", DONE),
    ("S5 并行分片", "分片规划", "模板匹配(Phase4)",
     "已实现：TEMPLATES → _build_spec_from_template：14 角色 → placement 映射"
     "（ndim 感知，D-08）",
     ALL, 1, "P0", "05 §3.5 / §12.2", DONE),
    ("S5 并行分片", "分片规划", "链式传播(Phase5)",
     "已实现：填充缺省 in_src + 相邻契约校验 + terminal 标记（名字无关单 entry 配对）",
     ALL, 1.5, "P0", "05 §3.6.5 / §12.3", DONE),
    ("S5 并行分片", "分片规划", "特殊参数处理(Phase6)",
     "已实现：SPECIAL_HANDLERS 注册表机制（已交付 gated_delta_tp_shard）；融合 QKV 走 "
     "FUSED_QKV 角色模板；DeepSeek MLA 经内置 ARCH_OVERRIDES（D-13），无需 SpecialHandler",
     f"{QW} / {DS}", 1, "P0", "05 §6.4.6/§12.2 D-13", DONE),
    ("S5 并行分片", "分片执行", "参数 DTensor 化",
     "已实现：distribute_tensor 按 spec.params 切分（含 EP Shard(0) 统一路径），支持 meta tensor",
     ALL, 0.5, "P0", "05 §4.2", DONE),
    ("S5 并行分片", "分片执行", "预编译通信计划",
     "已实现：precompiled_boundary.py PrecompiledBoundary：in_src→in_dst / out_src→out_dst "
     "编译为静态 RedistOp 序列，运行时零推导；identity 维跳过",
     ALL, 2, "P0", "05 §4.3", DONE),
    ("S5 并行分片", "分片执行", "生产模式前向包装",
     "已实现：boundary 入口 redistribute → local tensor 计算 → boundary 出口（参数已永久解包）",
     ALL, 0.5, "P0", "05 §4.4.1", DONE),
    ("S5 并行分片", "分片执行", "校验模式前向包装",
     "已实现：DTensor 全程传播 + out_src（核心）/out_dst（terminal）校验 + "
     "PlacementMismatchError 报告",
     ALL, 1, "P0", "05 §5", DONE),
    ("S5 并行分片", "分片执行", "CP attention 包装",
     "已实现：双模式同一 all-gather wrapper（D-01''，ring 已否决）：cp_utils.flex_cp_allgather "
     "K/V all-gather + offset-aware causal mask（D-04）",
     ALL, 1, "P0", "05 §4.4.2 / §12", DONE),
    ("S5 并行分片", "分片执行", "MoE local region 包装",
     "已实现：local_region.py（D-03' 泛化骨架）：boundary 入口 → local all-to-all → "
     "按声明 out_src 重包装，双模式共用",
     f"{DS} / {QW}", 0.5, "P0", "05 §4.4.3 / §12", DONE),
    ("S5 并行分片", "分片执行", "HF 原生 MoE EP 直通(D-09)",
     "已实现：planner stacked 元数据（_ep_stack/_moe_router）+ Phase A per-expert 参数堆叠为 "
     "[E,...] + wrapper 注入 _hf_native_ep_compute（router adapter + 本地 SwiGLU + 加权聚合），"
     "HF 单卡 MoE 脚本 EP>1 零改动",
     f"{DS} / {QW} / {GLM}", 3, "P0", "05 §6.4.7", DONE),
    ("S5 并行分片", "分片执行", "EP all_to_all 后端分派",
     "已实现：NCCL/HCCL 不等长 all_to_all（零填充）+ gloo pad-to-max all_to_all_single"
     "（测试路径），_EPAllToAllUneven/_EPAllToAllPadded autograd 双实现 + _ep_all_to_all 统一入口",
     f"{DS} / {QW}", 1, "P0", "05 §6.4.7", DONE),
    ("S5 并行分片", "分片执行", "MoE router adapter 注册表",
     "已实现：MOE_ROUTER_ADAPTERS 8 键（default/qwen3moe/qwen3_moe/mixtral/deepseekv3/"
     "deepseek_v3/glm4moe/glm4_moe）3 实现（softmax-topk / topk-router-module / sigmoid-group），"
     "未注册 arch 回落 default",
     f"{DS} / {QW} / {GLM}", 0.5, "P0", "05 §6.4.7", DONE),
    ("S5 并行分片", "分片执行", "TP-extend-EP 扩展专家并行(D-10)",
     "已实现：MindSpeed TP 扩展 EP / Megatron etp=1+ep 跨 TP 同构：ep_size 即扩展 EP 组大小"
     "（无单独 etp 配置），校验 ep_size ≤ dp_replicate×dp_cp×tp 且整除、"
     "num_experts % ep_size == 0。MoE 边界 SP-in identity + 全 dense 区域重分区 "
     "(edp,ep) + expert 权重仅 {EP:S0}（完整 expert）+ region 内 a2a（无 AG/RS）。"
     "D-11：HF 2025 batched 布局（experts.gate_up_proj [E,2I,H]）免堆叠直通",
     f"{DS} / {QW}", 5, "P1", "05 §6.4.8 / 06 §4.5.1", DONE),
    ("S5 并行分片", "分片执行", "派生 expert mesh 构建(D-10)",
     "已实现：sharding_applier._build_expert_mesh——主 mesh 无 EP 轴，apply 期将全 dense 区域 "
     "flatten 重分区为派生 expert mesh (edp,ep)，ep_size 整除校验",
     ALL, 0.5, "P0", "05 §6.4.8 / 06 §4.5.1", DONE),
    ("S5 并行分片", "分片执行", "vocab 并行 embedding 包装",
     "已实现：D-02 Megatron 风格 masked embedding（解包后 vocab mask 显式重建），仅 production 注入",
     ALL, 0.5, "P0", "05 §12.2 D-02", DONE),
    ("S5 并行分片", "分片执行", "双模式等价验证",
     "testing/grad_equiv.py 工具已交付（梯度等价 rtol=1e-3 + simulate_tp_replicate_grad_sync 旁路），"
     "TP 组合用例测绿；TP×CP/TP×EP 全组合输出等价覆盖与 FSDP2 集成（D-12 接线后）待补",
     ALL, 1, "P0", "05 §5.5 / dev_plan S5", PART),

    # ── S6 权重加载 ─────────────────────────────────────────────────────
    ("S6 权重加载", "键名映射", "权重键名映射",
     "预定义映射 + 模型自定义映射（如 _fp32_params.A_log ↔ A_log、MLA 参数重命名）",
     ALL, 1, "P0", "04 §5.3", TODO),
    ("S6 权重加载", "键名映射", "MoE/VLM 权重适配",
     "DeepSeek/Qwen3.5 MoE expert 合并拆分 + GLM-Image VLM 层级映射",
     f"{DS} / {QW} / {GLM}", 1, "P0", "04 §5.4", TODO),
    ("S6 权重加载", "并行加载", "并行权重加载",
     "所有 rank 独立读 safetensors，无需跨 rank 通信",
     ALL, 0.5, "P0", "04", TODO),
    ("S6 权重加载", "并行加载", "按 placement 分发",
     "完整权重写入已分片模型：自动按 placement 切分，保持原始 dtype",
     ALL, 0.5, "P0", "04", TODO),
    ("S6 权重加载", "验证", "加载往返测试",
     "save→load 一致 + 跨 TP 配置加载",
     ALL, 0.5, "P0", "04", TODO),

    # ── S7 训练组件 ─────────────────────────────────────────────────────
    ("S7 训练组件", "优化器", "优化器配置与构建",
     "AdamW 等配置声明 + factory escape hatch + 归一化入口",
     ALL, 1, "P0", "03", TODO),
    ("S7 训练组件", "优化器", "优化器实例化",
     "decay/no_decay 分组 + model.parts 遍历 + 每 part 独立优化器，"
     "返回 list[Optimizer]（canonical，nemo_automodel 惯例；04 OptimizerState 同步接受 list）",
     ALL, 0.5, "P0", "03", TODO),
    ("S7 训练组件", "调度器", "学习率调度器",
     "warmup + cosine decay，与 checkpoint 兼容（ratio-based wrapper）",
     ALL, 1, "P0", "03", TODO),
    ("S7 训练组件", "损失函数", "损失函数构建",
     "标准 CE + FusedLinearCrossEntropy + loss_parallel（vocab 并行 CE），"
     "自动选择 hidden_states/logits 路径",
     ALL, 1, "P0", "03 / 05 §3.5 lm_head", TODO),

    # ── S8 训练循环 ─────────────────────────────────────────────────────
    ("S8 训练循环", "流程编排", "有状态组件注册",
     "model/optimizer/scheduler/dataloader/rng 自动注册到状态追踪",
     ALL, 0.5, "P0", "03", TODO),
    ("S8 训练循环", "流程编排", "组件构建编排",
     "setup() 按依赖顺序构建全部训练组件（typed .build / untyped .instantiate）",
     ALL, 1, "P0", "03", TODO),
    ("S8 训练循环", "流程编排", "训练步进管理",
     "step/epoch 推进、验证/checkpoint/日志触发、梯度累积分组、SIGTERM 优雅退出",
     ALL, 1, "P0", "03", TODO),
    ("S8 训练循环", "流程编排", "训练循环执行",
     "epoch 迭代 → 训练步循环 → 定期验证 → 定期 checkpoint → GC → 安全退出",
     ALL, 0.5, "P0", "03", TODO),
    ("S8 训练循环", "训练步", "训练步执行(三阶段)",
     "统计全局 token 数 → 梯度累积循环（前后向）→ 梯度裁剪与参数更新 + scheduler 步进",
     ALL, 0.5, "P0", "03", TODO),
    ("S8 训练循环", "训练步", "前后向执行",
     "batch 上卡 → CP 上下文（shard_batch_for_cp）→ forward（PrecompiledBoundary）→ "
     "loss → backward → FSDP2 梯度同步（DTENSOR_UNIFIED / tp_grad_info 旁路，D-12 二选一）",
     ALL, 0.5, "P0", "03 / 05 §6.7", TODO),
    ("S8 训练循环", "训练步", "验证评估",
     "定期验证集前向 + 指标汇总（各 rank 归约）",
     ALL, 0.5, "P1", "03", TODO),

    # ── S9 检查点与持久化 ───────────────────────────────────────────────
    ("S9 持久化", "检查点", "检查点配置与管理器",
     "保存路径/间隔/HF 导出开关/异步开关 + StorageWriter/Reader + Addons 注册",
     ALL, 1, "P0", "04", TODO),
    ("S9 持久化", "检查点", "模型权重保存",
     "5 阶段：ModelState → Adapter.to_hf → index mapping → dcp.save → consolidate",
     ALL, 1, "P0", "04", TODO),
    ("S9 持久化", "检查点", "模型权重恢复",
     "3 路径：MoE tensor merging / Safetensors fast path / DCP resume + key_mapping",
     ALL, 1, "P0", "04", TODO),
    ("S9 持久化", "检查点", "MoE stacked 参数 key 映射",
     "D-09 堆叠后 experts.{proj} ↔ HF per-expert experts.{i}.{proj}.weight："
     "HF 导出时 unstack 拆回（arch 注册映射表）、init 加载零转换（先加载后 apply）、"
     "DCP resume 直存 stacked key",
     f"{DS} / {QW} / {GLM}", 1, "P0", "04 §7.6 / 05 §6.4.7", TODO),
    ("S9 持久化", "检查点", "HF 格式导出/加载",
     "单一 path 参数，内部解析 root_dir+model_name，本地缓存 fallback",
     ALL, 0.5, "P1", "04", TODO),
    ("S9 持久化", "状态管理", "模型/优化器状态管理",
     "DCP 兼容的 ModelState/OptimizerState（接受 list[Optimizer]），tied weights + PEFT 处理",
     ALL, 0.5, "P0", "04", TODO),
    ("S9 持久化", "状态管理", "分布式元数据记录",
     "记录 DTensor placement 信息，用于调试审计和 ShardingPlan diff",
     ALL, 0.5, "P1", "04", TODO),
    ("S9 持久化", "状态管理", "异步保存与故障恢复",
     "dcp.async_save 后台写入 + DistributedSignalHandler + LATEST symlink 原子更新 + 断点续训",
     ALL, 1, "P0", "04", TODO),

    # ── S10 高级并行特性 ────────────────────────────────────────────────
    ("S10 高级并行", "流水线并行", "PP 层切分",
     "transformer 层按 PP size 均分到不同 GPU，管理跨 stage 激活传递",
     ALL, 1.5, "P1", "06", TODO),
    ("S10 高级并行", "流水线并行", "PP 调度策略",
     "GPipe（同步）/ 1F1B（异步交错）两种调度 + causal mask 预计算",
     ALL, 1.5, "P1", "06", TODO),
    ("S10 高级并行", "显存优化", "激活检查点",
     "checkpoint_wrapper 每层包裹，full/selective 开关，以计算换显存",
     ALL, 1, "P0", "06", TODO),
    ("S10 高级并行", "混合精度", "混合精度策略",
     "param_dtype/reduce_dtype/output_dtype 配置（bf16/fp16）",
     ALL, 0.5, "P0", "06", TODO),
    ("S10 高级并行", "混合精度", "特定模块精度隔离",
     "指定模块（norm/router 等）固定 fp32，不随全局 dtype 转换",
     f"{ALL}（{DS} router 敏感）", 0.5, "P1", "06", TODO),
    ("S10 高级并行", "验证", "高级并行 E2E 测试",
     "PP=2/4 loss 一致 + AC 显存降低 >30%",
     ALL, 1, "P1", "—", TODO),

    # ── S11 模型交付适配 ────────────────────────────────────────────────
    ("S11 模型交付", "DeepSeek V3/V4", "架构接入与注册",
     "基于 HuggingFace 接入：架构注册 + HFCheckpointingMixin + from_pretrained 全链路",
     DS, 1, "P0", "01 §12", TODO),
    ("S11 模型交付", "DeepSeek V3/V4", "MLA 注意力分片",
     "planner 级已实现：_DEEPSEEK_MLA_OVERRIDES 内置 ARCH_OVERRIDES（D-13：q_a/kv_a 下投影 → "
     "REPLICATED（latent 维不切），q_b/kv_b 上投影 → COLWISE（head 维），o_proj 仍 ROWWISE；"
     "architectures 与 model_type 双拼写注册，v2/v3 同构）；无端到端训练验证",
     DS, 2, "P0", "05 §3.6.1/§12.2 D-13", PART),
    ("S11 模型交付", "DeepSeek V3/V4", "细粒度 expert EP 分片",
     "EP 直通已实现：D-09/D-10 覆盖细粒度 routed experts（deepseekv3 sigmoid-group router "
     "adapter + 派生 expert mesh (edp,ep) + region 内 a2a），shared expert 走 SHARED_EXPERT 角色；"
     "DeepEP/UCCL-EP 后端 token dispatcher 未实现",
     DS, 2, "P0", "05 §6.4.7/§6.4.8", PART),
    ("S11 模型交付", "DeepSeek V3/V4", "MTP 模块适配",
     "Multi-Token Prediction 模块的分片与损失链路接入",
     DS, 1, "P1", "—", TODO),
    ("S11 模型交付", "DeepSeek V3/V4", "权重映射与 E2E 测试",
     "V3/V4 权重键名映射 + TP/EP 组合 100 步训练 + 输出与 HF 参考容差 1e-5",
     DS, 1.5, "P0", "04 §5.4", TODO),
    ("S11 模型交付", "Qwen3.5", "Dense 架构实现",
     "GatedDeltaNet 层 + MTP 特殊逻辑 + 架构注册",
     QW, 1.5, "P0", "01 §12", TODO),
    ("S11 模型交付", "Qwen3.5", "架构覆盖规则",
     "已实现：GatedDeltaNet SPECIAL 角色 + gated_delta_tp_shard SpecialHandler（按 SSM head 切分，"
     "含 a_log/dt_bias 模式映射），components/distributed 交付并测绿；模型级接入属上行条目",
     QW, 1, "P0", "05 §6.4.6", DONE),
    ("S11 模型交付", "Qwen3.5", "MoE 架构实现",
     "MoE expert 合并 + EP 分片已实现（D-09/D-10：qwen3moe router adapter + 堆叠/batched 直通 + "
     "派生 expert mesh）；FSDP2 下 expert 梯度同步待 D-12 集成接线",
     QW, 2, "P0", "05 §6.4", PART),
    ("S11 模型交付", "Qwen3.5", "E2E 测试",
     "Dense/MoE 两变体 TP=2/4 100 步 + 输出容差 1e-5",
     QW, 1, "P0", "—", TODO),
    ("S11 模型交付", "GLM-Image", "VLM 架构接入",
     "基于 HuggingFace：vision encoder + LLM 的 ForImageTextToText 接入与架构注册",
     GLM, 1.5, "P0", "01 §6.1", TODO),
    ("S11 模型交付", "GLM-Image", "mRoPE 位置编码适配",
     "多模态旋转位置编码的 reshape 边界处理（R7：Shard(N) 无法表达逻辑轴——"
     "用户显式声明 reshape 后模块的 in_src）",
     GLM, 1, "P0", "05 §3.6.5 局限性", TODO),
    ("S11 模型交付", "GLM-Image", "权重层级映射",
     "VLM 视觉塔/投影层/语言模型的权重键名层级映射",
     GLM, 1, "P0", "04 §5.4", TODO),
    ("S11 模型交付", "GLM-Image", "E2E 测试",
     "图文混合 batch TP=2 训练 100 步 + 与 HF 参考一致",
     GLM, 1, "P0", "—", TODO),
    ("S11 模型交付", "Pangu(自研)", "自研模型类实现",
     "走内置自研路径（非 HF AutoModel）：模型结构实现 + 内置模型注册表接入",
     PG, 2.5, "P0", "01 §5/§7", TODO),
    ("S11 模型交付", "Pangu(自研)", "权重与分词器适配",
     "Pangu 私有权重格式与词表的加载适配（键名映射 + safetensors 加载路径）",
     PG, 1, "P0", "04", TODO),
    ("S11 模型交付", "Pangu(自研)", "分片规则注册",
     "Pangu 命名的 ARCH_OVERRIDES + 特殊参数 SpecialHandler 注册",
     PG, 1, "P0", "05 §8.3", TODO),
    ("S11 模型交付", "Pangu(自研)", "E2E 测试",
     "Pangu TP=2/4 训练 100 步 + 与参考实现对拍",
     PG, 1, "P0", "—", TODO),

    # ── S12 CLI 与监控 ──────────────────────────────────────────────────
    ("S12 CLI与监控", "命令行", "命令行入口",
     "hyper-parallel config.yaml --train.max_steps 500 风格 CLI",
     ALL, 1, "P1", "—", TODO),
    ("S12 CLI与监控", "日志", "控制台与文件日志",
     "Rank 0 控制台输出 + JSONL 文件记录",
     ALL, 0.5, "P1", "—", TODO),
    ("S12 CLI与监控", "日志", "远程日志上报",
     "WandB + MLflow 对接",
     ALL, 0.5, "P2", "—", TODO),
    ("S12 CLI与监控", "指标", "训练指标计算",
     "loss / grad_norm / lr / tokens/sec / MFU",
     ALL, 0.5, "P1", "—", TODO),
    ("S12 CLI与监控", "模板", "配置模板",
     "4 套 YAML 参考配置（DeepSeek MoE / Qwen3.5 / GLM-Image VLM / Pangu 多节点）",
     ALL, 0.5, "P1", "—", TODO),

    # ── S13 质量保障 ────────────────────────────────────────────────────
    ("S13 质量保障", "等价性验证", "零拷贝正确性验证",
     "已实现：DTensor._local_tensor 修改与全局视图同步（data_ptr 共享断言），"
     "tests/components/distributed 覆盖",
     ALL, 0.5, "P0", "05 §4.4", DONE),
    ("S13 质量保障", "等价性验证", "校验/生产模式等价性",
     "两模式输出/梯度等价用例已交付（test_dist_s5_mode_equiv / test_dist_s5_grad_equiv）；"
     "FSDP2 梯度同步集成测试待 D-12 接线后补（当前 grad_equiv.py 模拟旁路）",
     ALL, 1, "P0", "05 §5.5", PART),
    ("S13 质量保障", "等价性验证", "分片通信组合覆盖",
     "已实现：所有 in_src→in_dst 组合的 PrecompiledBoundary 正确性（identity/all_gather/"
     "reduce_scatter/all_reduce/redistribute 5 种 collective × 2 模式）",
     ALL, 0.5, "P0", "05 §4.3", DONE),
    ("S13 质量保障", "模型兼容", "交付模型兼容性测试",
     "DeepSeek V3/V4、Qwen3.5、GLM-Image、Pangu from_pretrained 推理与参考一致",
     f"{DS} / {QW} / {GLM} / {PG}", 1, "P0", "—", TODO),
    ("S13 质量保障", "持久化验证", "Checkpoint 往返测试",
     "save→load 一致 + 跨 TP 配置加载 + 断点续训 loss 连续",
     ALL, 0.5, "P0", "04", TODO),
    ("S13 质量保障", "端到端", "端到端训练验收",
     "交付模型 8 GPU 1000 步 loss 正常下降 + 断点续训",
     ALL, 1, "P0", "—", TODO),
    ("S13 质量保障", "文档", "迁移指南",
     "旧配置 → 新配置映射 + API 变化对照",
     ALL, 0.5, "P2", "—", TODO),
    ("S13 质量保障", "文档", "快速入门与自定义模型指南",
     "5 步跑通第一个训练 + 新模型添加 ShardingTemplate/ARCH_OVERRIDES 指南",
     ALL, 1, "P1", "—", TODO),
    ("S13 质量保障", "文档", "配置参考手册与设计文档校对",
     "YAML 字段完整说明 + 设计文档最终更新（含实施校准回写）",
     ALL, 0.5, "P1", "—", TODO),
]

FLOW = [
    ("S0 前置依赖", "TP 梯度同步集成（D-12 二选一）+ 版本钉板 + 交付模型可获得性确认",
     "解锁生产模式 TP 梯度同步；不阻塞正确性（可降级校验模式）"),
    ("S1 启动与配置", "CLI 读入 YAML → parse_training_args → TrainerConfig 强类型配置【已实现】", "用户写一个 YAML 即可启动训练"),
    ("S2 分布式环境", "并行度声明 → DeviceMesh（主 mesh 无 EP 轴，D-10）→ FSDP2 包裹 + TP 梯度元数据", "多卡通信拓扑与数据并行就绪"),
    ("S3 数据供给", "tokenizer → dataset → packing → sampler → collate → DataLoader", "数据按 DP/CP 契约切好喂给模型"),
    ("S4 模型构建", "模型注册 → from_pretrained → meta 空壳 → PEFT/freeze", "模型结构就绪（零显存）"),
    ("S5 并行分片", "ShardingPlanner 推导 → apply_sharding_plan 应用（双模式）", "每个参数怎么切自动推导，通信预编译【已实现，282 用例全绿】"),
    ("S6 权重加载", "键名映射 → 并行加载 → 按 placement 分发", "预训练权重落到分片后的模型"),
    ("S7 训练组件", "优化器（list[Optimizer]）+ LR 调度器 + 损失函数", "训练步所需组件就绪"),
    ("S8 训练循环", "setup 编排 → 训练步（前后向/梯度累积/裁剪/更新）→ 定期验证", "训练主流程"),
    ("S9 持久化", "checkpoint 保存/恢复 + HF 导出 + 异步保存 + 断点续训", "训练状态可保存可恢复"),
    ("S10 高级并行", "PP 切分与调度 + 激活检查点 + 混合精度", "更大模型/更长序列"),
    ("S11 模型交付", "DeepSeek V3/V4 + Qwen3.5 + GLM-Image + Pangu(自研) 适配", "4 个交付模型全部可训练"),
    ("S12 CLI与监控", "命令行入口 + 日志 + 指标 + 配置模板", "训练可启动可观测"),
    ("S13 质量保障", "双模式等价 + 组合覆盖 + 4 模型兼容 + E2E 验收 + 文档", "交付质量兜底"),
]

MODELS = [
    ("DeepSeek V3/V4", "HuggingFace", "MoE（细粒度 expert + shared expert）+ MLA 注意力 + MTP",
     "MLA 内置 ARCH_OVERRIDES（D-13，planner 级已实现）；EP 直通（D-09/D-10 已实现）；"
     "router 精度隔离；V3/V4 权重键名映射与 E2E 验证", "7.5"),
    ("Qwen3.5", "HuggingFace", "Dense（GatedDeltaNet + MTP）与 MoE 两变体",
     "GatedDeltaNet SPECIAL 分片（已实现）；MoE EP 直通（已实现）；模型注册/FSDP2 集成与两变体 E2E", "5.5"),
    ("GLM-Image", "HuggingFace", "VLM（vision encoder + LLM），ForImageTextToText",
     "mRoPE reshape 边界显式声明；NEAT 图像打包；VLM 权重层级映射", "4.5"),
    ("Pangu", "自研（内置路径，非 HF）", "自研架构",
     "自研模型类实现与内置注册；私有权重/词表适配；ARCH_OVERRIDES 分片规则注册", "5.5"),
]

STATUS_ORDER = (DONE, PART, TODO)


def _status_counts(rows):
    return {s: sum(1 for r in rows if r[8] == s) for s in STATUS_ORDER}


def _style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="2F5597")
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _thin_border():
    side = Side(style="thin", color="B0B0B0")
    return Border(left=side, right=side, top=side, bottom=side)


def gen_md():
    total = sum(r[5] for r in ROWS)
    counts = _status_counts(ROWS)
    lines = [
        "# Hyper-Parallel 需求分解（端到端训练流程版）",
        "",
        "> 本文档与 `requirements.xlsx` / `requirements.csv` 同源生成"
        "（2026-07-22 更新：状态列按 components/distributed 实现实况标注）。",
        "> 组织方式：**按大模型训练端到端流程（S0→S13）列出需求**，每个需求点按"
        "「模块 + 功能」划分，可直接交由程序员开发。",
        "> 交付模型：**DeepSeek V3/V4、Qwen3.5、GLM-Image**（基于 HuggingFace 接入）、"
        "**Pangu**（自研，走内置模型路径）。",
        f"> 实现状态：**已实现 {counts[DONE]} / 部分实现 {counts[PART]} / 未实现 {counts[TODO]}**"
        "（components/distributed 已落地，tests/components/distributed 282 用例全绿）。",
        "",
        "## 端到端流程总览",
        "",
        "| 流程阶段 | 阶段内容 | 阶段目标 |",
        "|---------|---------|---------|",
    ]
    for stage, content, goal in FLOW:
        lines.append(f"| {stage} | {content} | {goal} |")
    lines += [
        "",
        "> 流程说明：S0→S1→S2 为启动期；S3 与 S4→S5→S6 可并行准备；S7→S8 进入训练主循环；",
        "> S9 贯穿训练周期；S10 按需开启；S11 与各基础模块并行开发；S12/S13 覆盖全流程。",
        "",
        "## 需求总表",
        "",
    ]

    stage_order = []
    for r in ROWS:
        if r[0] not in stage_order:
            stage_order.append(r[0])

    idx = 0
    for stage in stage_order:
        rows = [r for r in ROWS if r[0] == stage]
        subtotal = sum(r[5] for r in rows)
        lines.append(f"### {stage}（小计 {subtotal} 人·日）")
        lines.append("")
        lines.append("| # | 模块 | 功能/需求点 | 需求描述 | 适用模型 | 工时 | 优先级 | 设计文档 | 状态 |")
        lines.append("|--:|------|------------|---------|---------|:----:|:------:|---------|:----:|")
        for _, module, point, desc, model, effort, prio, doc, status in rows:
            idx += 1
            desc = desc.replace("|", "\\|")
            lines.append(
                f"| {idx} | {module} | {point} | {desc} | {model} | {effort} | {prio} | {doc} | {status} |"
            )
        lines.append("")

    lines += [
        "## 模型交付",
        "",
        "| 交付模型 | 接入方式 | 架构特点 | 专项适配需求 | 工时小计 |",
        "|---------|---------|---------|-------------|:-------:|",
    ]
    for name, way, feat, work, effort in MODELS:
        lines.append(f"| {name} | {way} | {feat} | {work} | {effort} |")
    lines += [
        "",
        "## 汇总",
        "",
        "| 流程阶段 | 需求点数 | 工时小计（人·日） | 已实现 | 部分实现 | 未实现 |",
        "|---------|:-------:|:----------------:|:-----:|:-------:|:-----:|",
    ]
    for stage in stage_order:
        rows = [r for r in ROWS if r[0] == stage]
        n = len(rows)
        e = sum(r[5] for r in rows)
        c = _status_counts(rows)
        lines.append(f"| {stage} | {n} | {e} | {c[DONE]} | {c[PART]} | {c[TODO]} |")
    lines.append(
        f"| **合计** | **{len(ROWS)}** | **{total}** | "
        f"**{counts[DONE]}** | **{counts[PART]}** | **{counts[TODO]}** |"
    )
    lines.append("")
    return "\n".join(lines)


def gen_csv(docs_dir):
    with open(f"{docs_dir}/requirements.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["序号", "流程阶段", "模块", "功能/需求点", "需求描述",
                    "适用模型", "工时(人·日)", "优先级", "设计文档", "状态"])
        for i, row in enumerate(ROWS, 1):
            w.writerow([i, *row])


def build():
    wb = Workbook()

    # ── Sheet 1: 需求总表 ──
    ws = wb.active
    ws.title = "需求总表"
    headers = ["序号", "流程阶段", "模块", "功能/需求点", "需求描述",
               "适用模型", "工时(人·日)", "优先级", "设计文档", "状态"]
    ws.append(headers)
    _style_header(ws)

    stage_fill = {
        "S0": "F8CBAD", "S1": "FFE699", "S2": "C6E0B4", "S3": "BDD7EE",
        "S4": "D9E1F2", "S5": "B4C6E7", "S6": "A9D08E", "S7": "FFD966",
        "S8": "F4B084", "S9": "D0CECE", "S10": "E2EFDA", "S11": "FBE2D5",
        "S12": "DDEBF7", "S13": "E4DFEC",
    }
    done_fill = PatternFill("solid", fgColor="E2EFDA")   # 已实现：浅绿整行
    part_fill = PatternFill("solid", fgColor="FFF2CC")   # 部分实现：浅黄（仅状态格）
    border = _thin_border()
    for i, row in enumerate(ROWS, 1):
        stage, status = row[0], row[8]
        ws.append([i, *row])
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True,
                                       horizontal="center" if c in (1, 2, 6, 7, 8, 10) else "left")
        if status == DONE:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = done_fill
        elif status == PART:
            ws.cell(row=r, column=10).fill = part_fill
        key = stage.split()[0]
        if key in stage_fill:
            ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=stage_fill[key])
    widths = [6, 13, 18, 24, 60, 16, 11, 8, 18, 10]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"
    # 合计行
    total = sum(r[5] for r in ROWS)
    counts = _status_counts(ROWS)
    ws.append(["", "合计", "", "",
               f"共 {len(ROWS)} 个需求点", "", total, "", "",
               f"已实现 {counts[DONE]} / 部分 {counts[PART]} / 未实现 {counts[TODO]}"])
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.font = Font(bold=True)
        cell.border = border

    # ── Sheet 2: 流程总览 ──
    ws2 = wb.create_sheet("流程总览")
    ws2.append(["流程阶段", "阶段内容（端到端训练流水线）", "阶段目标"])
    _style_header(ws2)
    for stage, content, goal in FLOW:
        ws2.append([stage, content, goal])
        for c in range(1, 4):
            ws2.cell(row=ws2.max_row, column=c).border = border
            ws2.cell(row=ws2.max_row, column=c).alignment = Alignment(
                vertical="center", wrap_text=True)
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 70
    ws2.column_dimensions["C"].width = 40
    ws2.freeze_panes = "A2"
    note_row = ws2.max_row + 2
    ws2.cell(row=note_row, column=1,
             value="流程说明：S0→S1→S2 为启动期；S3 与 S4→S5→S6 可并行准备；"
                   "S7→S8 进入训练主循环；S9 贯穿训练周期；S10 按需开启；"
                   "S11 与各基础模块并行开发；S12/S13 覆盖全流程。"
                   "S5 并行分片引擎已实现（components/distributed，282 用例全绿）。")
    ws2.cell(row=note_row, column=1).font = Font(italic=True, color="808080")
    ws2.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)

    # ── Sheet 3: 模型交付 ──
    ws3 = wb.create_sheet("模型交付")
    ws3.append(["交付模型", "接入方式", "架构特点", "专项适配需求", "工时小计(人·日)"])
    _style_header(ws3)
    for name, way, feat, work, effort in MODELS:
        ws3.append([name, way, feat, work, effort])
        for c in range(1, 6):
            ws3.cell(row=ws3.max_row, column=c).border = border
            ws3.cell(row=ws3.max_row, column=c).alignment = Alignment(
                vertical="center", wrap_text=True,
                horizontal="center" if c == 5 else "left")
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 24
    ws3.column_dimensions["C"].width = 40
    ws3.column_dimensions["D"].width = 60
    ws3.column_dimensions["E"].width = 14
    ws3.freeze_panes = "A2"

    # ── Sheet 4: 汇总 ──
    ws4 = wb.create_sheet("汇总")
    ws4.append(["流程阶段", "需求点数", "工时小计(人·日)", "已实现", "部分实现", "未实现"])
    _style_header(ws4)
    stage_order = []
    for r in ROWS:
        if r[0] not in stage_order:
            stage_order.append(r[0])
    for stage in stage_order:
        rows = [r for r in ROWS if r[0] == stage]
        c = _status_counts(rows)
        ws4.append([stage, len(rows), sum(r[5] for r in rows),
                    c[DONE], c[PART], c[TODO]])
        for col in range(1, 7):
            ws4.cell(row=ws4.max_row, column=col).border = border
            ws4.cell(row=ws4.max_row, column=col).alignment = Alignment(horizontal="center")
    counts = _status_counts(ROWS)
    ws4.append(["合计", len(ROWS), sum(r[5] for r in ROWS),
                counts[DONE], counts[PART], counts[TODO]])
    for col in range(1, 7):
        ws4.cell(row=ws4.max_row, column=col).font = Font(bold=True)
        ws4.cell(row=ws4.max_row, column=col).border = border
        ws4.cell(row=ws4.max_row, column=col).alignment = Alignment(horizontal="center")
    ws4.column_dimensions["A"].width = 16
    for col_letter in ("B", "C", "D", "E", "F"):
        ws4.column_dimensions[col_letter].width = 14

    import os
    docs_dir = os.path.dirname(os.path.abspath(__file__))
    wb.save(os.path.join(docs_dir, "requirements.xlsx"))
    with open(os.path.join(docs_dir, "requirements.md"), "w", encoding="utf-8") as f:
        f.write(gen_md())
    gen_csv(docs_dir)
    counts = _status_counts(ROWS)
    print(f"rows={len(ROWS)}, total={sum(r[5] for r in ROWS)}, "
          f"done={counts[DONE]}, part={counts[PART]}, todo={counts[TODO]} -> xlsx/md/csv")


if __name__ == "__main__":
    build()
